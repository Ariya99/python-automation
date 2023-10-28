import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

def is_valid_csv_filename(filename):
    # Check if the filename ends with .csv (case insensitive)
    return filename.lower().endswith(".csv")

# Get the filenames for the "failed" and "all_operations" CSV files
while True:
    failed_file = input("Enter the filename for the 'failed' CSV file: ")
    if is_valid_csv_filename(failed_file) and os.path.exists(failed_file):
        break
    else:
        print("Invalid filename. Please provide a valid .csv file.")

while True:
    all_operations_file = input("Enter the filename for the 'all_operations' CSV file: ")
    if is_valid_csv_filename(all_operations_file) and os.path.exists(all_operations_file):
        break
    else:
        print("Invalid filename. Please provide a valid .csv file.")

# Read the data from the CSV files
df_failed = pd.read_csv(failed_file, delimiter=',')
df_all = pd.read_csv(all_operations_file, delimiter=',')

# Merge the dataframes based on the "OPERATION_NAME" column
merged_df = df_all.merge(df_failed, on="OPERATION_NAME", how="left", suffixes=("_all", "_failed"))

# Fill missing values in the "OPERATION_COUNT_failed" column with 0
merged_df["OPERATION_COUNT_failed"].fillna(0, inplace=True)

# Calculate the percentage of failing for each operation
merged_df["Failure Percentage"] = (merged_df["OPERATION_COUNT_failed"] / merged_df["OPERATION_COUNT_all"]) * 100

# Get the date input from the user
date_input = input("Enter the date (e.g., 2oct_10oct): ")

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Add the header row
header = ["OPERATION_NAME", "OPERATION_COUNT_all", "OPERATION_COUNT_failed", "Failure Percentage"]
ws.append(header)

# Write the data to the Excel worksheet
for r_idx, row in merged_df.iterrows():
    ws.append(row.tolist())

# Apply conditional formatting based on the "Failure Percentage" column
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):  # 4 corresponds to the "Failure Percentage" column
    for cell in row:
        percentage = cell.value
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        if percentage > 10 and percentage <= 33:
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        elif percentage > 33:
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        cell.fill = fill

# Save the Excel workbook with the provided file name including the date
output_file = f"Syncprocess_health_report_{date_input}.xlsx"
wb.save(output_file)

# Add a message indicating that the data has been combined and saved to the file
print("Data combined and saved to file:", output_file)

