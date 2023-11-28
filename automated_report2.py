import pandas as pd
from openpyxl.styles import PatternFill

# Prompt the user to input the file names for failed, executed, and all status CSV files
failed_filename = input("Enter the name of the failed operations CSV file: ")
executed_filename = input("Enter the name of the executed operations CSV file: ")
all_status_filename = input("Enter the name of the all status operations CSV file: ")

# Check if the file names have .csv or .CSV extension and add if missing
failed_filename = failed_filename if failed_filename.lower().endswith('.csv') else failed_filename + '.csv'
executed_filename = executed_filename if executed_filename.lower().endswith('.csv') else executed_filename + '.csv'
all_status_filename = all_status_filename if all_status_filename.lower().endswith('.csv') else all_status_filename + '.csv'

# Read the csv files and store them in dataframes
df_failed = pd.read_csv(failed_filename)
df_executed = pd.read_csv(executed_filename)
df_all_status = pd.read_csv(all_status_filename)

# Concatenate the dataframes vertically
df_concat = pd.concat([df_failed, df_executed, df_all_status])

# Create a new dataframe with unique operation names
df_unique = pd.DataFrame({'Operation_Name': df_concat['OPERATION_NAME'].unique()})

# Prompt the user to input the output Excel file name
output_filename = input("Enter the output Excel file name: ")
output_filename = 'SyncHealth_Report_' + output_filename + '.xlsx'

# Initialize columns for executed counts, all status counts, failed counts, and failing rate
df_unique['Executed_Count'] = 0
df_unique['All_Status_Count'] = 0
df_unique['Failed_Count'] = 0
df_unique['Failing_Rate'] = 0.0

# Iterate over each operation name
for index, row in df_unique.iterrows():
    operation_name = row['Operation_Name']
    
    # Find the executed count for the operation name in the executed file
    executed_count = df_executed[df_executed['OPERATION_NAME'] == operation_name]['OPERATION_COUNT'].sum()
    
    # Update the corresponding column in the dataframe
    df_unique.loc[index, 'Executed_Count'] = executed_count
    
    # Find all status count, failed count, and failing rate for the operation name
    all_status_count = df_all_status[df_all_status['OPERATION_NAME'] == operation_name]['OPERATION_COUNT'].sum()
    failed_count = df_failed[df_failed['OPERATION_NAME'] == operation_name]['OPERATION_COUNT'].sum()
    failing_rate = (failed_count / all_status_count) * 100 if all_status_count != 0 else 0
    
    # Update the corresponding columns in the dataframe
    df_unique.loc[index, 'All_Status_Count'] = all_status_count
    df_unique.loc[index, 'Failed_Count'] = failed_count
    df_unique.loc[index, 'Failing_Rate'] = failing_rate

# Create the Excel writer object
writer = pd.ExcelWriter(output_filename, engine='openpyxl')

# Convert the dataframe to Excel
df_unique.to_excel(writer, index=False, sheet_name='Report')

# Get the workbook and worksheet objects
workbook = writer.book
worksheet = writer.sheets['Report']

# Define the cell fill colors
red_fill = PatternFill(fill_type='solid', fgColor='FF0000')
orange_fill = PatternFill(fill_type='solid', fgColor='FFA500')
green_fill = PatternFill(fill_type='solid', fgColor='00FF00')

# Get the range of cells in the failing rate column
failing_rate_column = 'E'
start_row = 2
end_row = len(df_unique) + 1
cell_range = f'{failing_rate_column}{start_row}:{failing_rate_column}{end_row}'

# Apply the cell fill colors based on the failing rate percentages
for cell in worksheet[cell_range]:
    for c in cell:
        value = c.value
        if value is not None:
            if value > 33:
                c.fill = red_fill
            elif 10 <= value <= 33:
                c.fill = orange_fill
            else:
                c.fill = green_fill

# Save the workbook
workbook.save(output_filename)
