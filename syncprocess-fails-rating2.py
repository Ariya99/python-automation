import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# File paths for the two CSV files
failed_file = "Report-2oct_10oct-Failed.csv"
all_operations_file = "Report-2oct_10oct-all_operations_status.csv"

# Read the data from the CSV files
df_failed = pd.read_csv(failed_file, delimiter=',')
df_all = pd.read_csv(all_operations_file, delimiter=',')

# Merge the dataframes based on the "OPERATION_NAME" column
merged_df = df_all.merge(df_failed, on="OPERATION_NAME", how="left", suffixes=("_all", "_failed"))

# Fill missing values in the "OPERATION_COUNT_failed" column with 0
merged_df["OPERATION_COUNT_failed"].fillna(0, inplace=True)

# Calculate the percentage of failing for each operation
merged_df["Failure Percentage"] = (merged_df["OPERATION_COUNT_failed"] / merged_df["OPERATION_COUNT_all"]) * 100

# Create a new Excel workbook and add a worksheet
wb = Workbook()
ws = wb.active

# Write the data to the Excel worksheet
for r_idx, row in merged_df.iterrows():
    ws.append(row.tolist())

# Remove the first row that contains "OPERATION_NAME" and "OPERATION_COUNT"
ws.delete_rows(1)

# Apply conditional formatting based on the "Failure Percentage" column
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=3, max_col=3):
    for cell in row:
        percentage = cell.value
        fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")  # Green
        if percentage > 10 and percentage <= 33:
            fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
        elif percentage > 33:
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # Red
        cell.fill = fill

# Save the Excel workbook with the provided file name
output_file = "Syncprocess_health_report_edited.xlsx"
wb.save(output_file)

